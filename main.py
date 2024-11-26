from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, NamedStyle, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
import pandas as pd
import time
import pyautogui
import os
import glob
import random
import xlwings as xw
import pdfkit
import os

# Obtener la ruta base del directorio donde está el script
base_dir = os.path.dirname(os.path.abspath(__file__))

# Leer el archivo Excel
df = pd.read_excel(os.path.join(base_dir, "Data", "Clientes.xlsx"))

print("Ruta excel de clientes", os.path.join(base_dir, "Data", "Clientes.xlsx"))

# Suposición de nombres de columnas
cuit_login_list = df['CUIT para ingresar'].tolist()
print(len(cuit_login_list))
cuit_represent_list = df['CUIT representado'].tolist()
password_list = df['Contraseña'].tolist()
posterior_list = df['Posterior'].tolist()
anterior_list = df['Anterior'].tolist()
clientes_list = df['Cliente'].tolist()

# Configuración de opciones de Chrome
options = Options()
options.add_argument("--start-maximized")

# Configurar preferencias de descarga
prefs = {
    "download.prompt_for_download": True,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
}
options.add_experimental_option("prefs", prefs)

# Inicializar driver
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

# Crear el archivo de resultados
resultados = []

def human_typing(element, text):
    for char in str(text):
        element.send_keys(char)
        time.sleep(random.uniform(0.05, 0.3))

def actualizar_excel(row_index, mensaje):
    """Actualiza la última columna del archivo Excel con un mensaje de error."""
    df.at[row_index, 'Error'] = mensaje
    df.to_excel(os.path.join(base_dir, "Data", "Clientes.xlsx"), index=False)

# Función para iniciar sesión
def iniciar_sesion(cuit_ingresar, password, row_index):
    """Inicia sesión en el sitio web con el CUIT y contraseña proporcionados."""
    try:
        driver.get('https://auth.afip.gob.ar/contribuyente_/login.xhtml')

        # Ingreso el CUIT
        element = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'F1:username')))
        element.clear()
        human_typing(element, cuit_ingresar)

        time.sleep(5)

        # Click en siguiente
        driver.find_element(By.ID, 'F1:btnSiguiente').click()

        # Verificar si el CUIT es incorrecto
        try:
            error_message = driver.find_element(By.ID, 'F1:msg').text
            if error_message == "Número de CUIL/CUIT incorrecto":
                actualizar_excel(row_index, "Número de CUIL/CUIT incorrecto")
                return False
        except:
            pass

        # Ingreso la contraseña
        element_pass = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'F1:password')))
        human_typing(element_pass, password)

        # Click en ingresar
        driver.find_element(By.ID, 'F1:btnIngresar').click()
        time.sleep(2)

        # Verificar si la contraseña es incorrecta
        try:
            error_message = driver.find_element(By.ID, 'F1:msg').text
            if error_message == "Clave o usuario incorrecto":
                actualizar_excel(row_index, "Clave o usuario incorrecto")
                return False
        except:
            pass

        return True
    except Exception as e:
        print(f"Error al iniciar sesión: {e}")
        actualizar_excel(row_index, "Error al iniciar sesión")
        return False

def ingresar_modulo(cuit_ingresar, password, row_index):
    """Ingresa al módulo específico del sistema de cuentas tributarias."""
    try:
        # Click en ver todos los módulos 
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.LINK_TEXT, "Ver todos"))).click()
        time.sleep(2)

        # Escribir en el buscador el módulo
        element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'buscadorInput')))
        human_typing(element, 'Sistema de Cuentas Tributarias') 
        time.sleep(2)

        # Clickear el módulo
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'rbt-menu-item-0'))).click()
        time.sleep(2)

        try:
            # Esperar y manejar el modal si aparece
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CLASS_NAME, 'modal-content')))
            modal = driver.find_element(By.CLASS_NAME, 'modal-content')
            if modal.is_displayed():
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Continuar"]'))).click()
        except:
            # No hacer nada si el modal no aparece
            pass

        # Cambiar de pestaña
        window_handles = driver.window_handles
        driver.switch_to.window(window_handles[-1])

        # Verificar mensaje de error de autenticación
        try:
            error_message = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.TAG_NAME, 'pre')))
            if error_message.text == "Ha ocurrido un error al autenticar, intente nuevamente.":
                actualizar_excel(row_index, "Error autenticacion")
                driver.refresh()
        except:
            pass

        # Verificar si es necesario iniciar sesion nuevamente
        try:
            element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'F1:username')))
            element.clear()
            human_typing(element, cuit_ingresar)
            driver.find_element(By.ID, 'F1:btnSiguiente').click()
            time.sleep(1)

            element_pass = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'F1:password')))
            human_typing(element_pass, password)
            time.sleep(5)
            driver.find_element(By.ID, 'F1:btnIngresar').click()
            time.sleep(1)

            actualizar_excel(row_index, "Error volver a iniciar sesion")
        except:
            pass

    except Exception as e:
        print(f"Error al ingresar al módulo: {e}")

def seleccionar_cuit_representado(cuit_representado):
    """Selecciona el CUIT representado en el sistema."""
    try:
        select_present = EC.presence_of_element_located((By.NAME, "$PropertySelection"))
        if WebDriverWait(driver, 10).until(select_present):
            current_selection = Select(driver.find_element(By.NAME, "$PropertySelection")).first_selected_option.text
            if current_selection != str(cuit_representado):
                select_element = Select(driver.find_element(By.NAME, "$PropertySelection"))
                select_element.select_by_visible_text(str(cuit_representado))
    except Exception:
        try:
            cuit_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'span.cuit')))
            cuit_text = cuit_element.text.replace('-', '')
            if cuit_text != str(cuit_representado):
                print(f"El CUIT ingresado no coincide con el CUIT representado: {cuit_representado}")
                return False
        except Exception as e:
            print(f"Error al verificar CUIT: {e}")
            return False
    # Esperar que el popup esté visible y hacer clic en el botón de cerrar por XPATH
    try:
    # Usamos el XPATH para localizar el botón de cerrar
        close_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, '//a[@href="#close" and @title="Cerrar"]'))
        )
        close_button.click()
        print("Popup cerrado exitosamente.")
    except Exception as e:
        print(f"Error al intentar cerrar el popup: {e}")
    return True

def cerrar_sesion():
    """Cierra la sesión actual."""
    try:
        driver.close()
        window_handles = driver.window_handles
        driver.switch_to.window(window_handles[0])
        driver.find_element(By.ID, "iconoChicoContribuyenteAFIP").click()
        driver.find_element(By.XPATH, '//*[@id="contBtnContribuyente"]/div[6]/button/div/div[2]').click()
        time.sleep(5)
    except Exception as e:
        print(f"Error al cerrar sesión: {e}")

def verificar_deuda(contador):
    print("Verificando deuda")

    if contador != 1:
        # Abrir menú
        xpath_menú = "/html/body/div[2]/div[1]/table/tbody/tr/td[1]/a/i"
        # Esperar hasta que el elemento sea clickeable
        element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_menú)))
        # Hacer clic en el elemento
        element.click()  
        time.sleep(1)

        print("abro menú")

    if contador == 1:
        # Clickear en cuenta corriente
        xpath_cuenta_corriente = "/html/body/div[3]/div[2]/div[1]/div[2]/table[3]/tbody/tr/td[1]/div"
        # Esperar hasta que el elemento sea clickeable
        element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_cuenta_corriente)))
        # Hacer clic en el elemento
        element.click()
        time.sleep(1)
        print("click cuenta corriente")

    # Clickear en "estado de cumplimiento"
    xpath_estado_cumplimiento = "/html/body/div[3]/div[2]/div[1]/div[2]/div[5]/div/table/tbody/tr[1]/td[2]/div"
    # Esperar hasta que el elemento sea clickeable
    element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_estado_cumplimiento)))
    # Hacer clic en el elemento
    element.click()
    time.sleep(1)
    print("click estado de cumplimiento")

    xpath_desplegar_opciones = "/html/body/div[3]/div[2]/div[2]/div[2]/div/form/div[2]/div[3]/div/span[1]/div/span/span[2]/div/span/span[1]/span/span[2]"
    # Esperar hasta que el elemento sea clickeable
    element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_desplegar_opciones)))
    # Hacer clic en el elemento
    element.click()
    time.sleep(1)
    print("click opciones")

    try:
        # Esperar a que la lista de opciones sea visible
        opcion_351_xpath = "//li[contains(@id, 'select2-chooser1-result') and contains(text(), '351 - CONTRIBUCIONES SEG. SOCIAL')]"
        # Esperar que el elemento esté clickeable
        opcion_351 = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, opcion_351_xpath)))
        # Hacer clic en la opción
        opcion_351.click()

        print("click en 351")
    except:
        actualizar_excel(indice, "No contiene 351")

    # Obtener la fecha actual
    fecha_actual = datetime.now()

    # Calcular el mes y el año del periodo
    mes_anterior = fecha_actual.month - 1
    if mes_anterior == 0:  # Si estamos en enero, el mes anterior es diciembre del año anterior
        mes_anterior = 12
        anio_anterior = fecha_actual.year - 1
    else:
        anio_anterior = fecha_actual.year

    # Formatear el periodo como "AAAAMM00"
    periodo = f"{anio_anterior}{mes_anterior:02}00"

    # XPath del campo
    campo_periodo_desde_xpath = "/html/body/div[3]/div[2]/div[2]/div[2]/div/form/div[2]/div[3]/div/span[2]/div/span/span[1]/div/div[2]/input"
    campo_periodo_hasta_xpath = "/html/body/div[3]/div[2]/div[2]/div[2]/div/form/div[2]/div[3]/div/span[2]/div/span/span[2]/div/div[2]/input"

    time.sleep(2)
    # Esperar a que el campo esté visible y escribir el periodo
    campo_periodo_desde = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, campo_periodo_desde_xpath)))
    campo_periodo_desde.clear()  # Limpiar el campo antes de escribir
    campo_periodo_desde.click()
    human_typing(campo_periodo_desde, periodo)

    time.sleep(1)
    # Esperar a que el campo esté visible y escribir el periodo
    campo_periodo_hasta = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, campo_periodo_hasta_xpath)))
    campo_periodo_hasta.clear()  # Limpiar el campo antes de escribir
    campo_periodo_hasta.click()
    human_typing(campo_periodo_hasta, periodo)

    print("periodos cargados")

    time.sleep(2)

    xpath_siguiente = "/html/body/div[3]/div[2]/div[2]/div[2]/div/form/div[2]/div[3]/div/span[6]/div/span/span/div"
    # Esperar que el elemento esté clickeable
    siguiente = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_siguiente)))
    # Hacer clic en la opción
    siguiente.click()

    print("click siguiente")

    time.sleep(5)

    xpath_deuda= "/html/body/div[3]/div[2]/div[2]/div[2]/div/form/div[2]/div[3]/div/span[2]/div/span/span[2]/div/span/span[2]/div/div"
    # Esperar a que el elemento sea visible y obtener su texto
    deuda_elemento = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, xpath_deuda)))
    deuda_valor = deuda_elemento.text  # Extraer el texto contenido en el elemento

    # Mostrar o guardar el valor de la deuda
    print("Valor de la deuda:", deuda_valor)

    # Entrar al módulo, realizar acciones e intentar obtener el valor de la deuda
    try:        
        # Si no es convertible a número, tratarlo como deuda inexistente
        try:
            deuda_valor_num = float(deuda_valor.replace(",", "").replace(".", "").replace(" ", ""))
        except ValueError:
            deuda_valor_num = 0.0
        
        # Determinar si hay deuda
        tiene_deuda = "Sí" if deuda_valor_num > 0 else "No"
    
    except Exception as e:
        # Si hay algún error, registrar como deuda no disponible
        deuda_valor_num = 0.0
        tiene_deuda = "No"
        print(f"Error obteniendo deuda para {cliente}: {e}")

    # Agregar los datos del cliente al registro
    datos_clientes.append({
        "Cliente": cliente,
        "Tiene Deuda": tiene_deuda,
        "Importe Deuda": deuda_valor_num
    })

def extraer_datos_nuevo(cuit_ingresar, cuit_representado, password, posterior, cliente, indice, contador):
    """Extrae datos para un nuevo usuario"""
    try:
        control_sesion = iniciar_sesion(cuit_ingresar, password, indice)
        if control_sesion:
            ingresar_modulo(cuit_ingresar, password, indice)
            # Esperar que el popup esté visible y hacer clic en el botón de cerrar por XPATH
            try:
                # Usamos el XPATH para localizar el botón de cerrar
                close_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//a[@href="#close" and @title="Cerrar"]'))
                )
                close_button.click()
                print("Popup cerrado exitosamente.")
            except Exception as e:
                print(f"Error al intentar cerrar el popup: {e}")
            if seleccionar_cuit_representado(cuit_representado):
                print("Verificar deuda")
                verificar_deuda(contador)
                if posterior == 0:
                    print("Cerrando sesión")
                    cerrar_sesion()
    except Exception as e:
        print(f"Error al extraer datos para el nuevo usuario: {e}")

def extraer_datos(cuit_representado, posterior, cliente, contador):
    """Extrae datos para un usuario existente."""
    try:
        if seleccionar_cuit_representado(cuit_representado):
            verificar_deuda(contador)
            if posterior == 0:
                cerrar_sesion()
    except Exception as e:
        print(f"Error al extraer datos: {e}")

# Función para convertir Excel a CSV utilizando xlwings
def excel_a_csv(input_folder, output_folder):
    for excel_file in glob.glob(os.path.join(input_folder, "*.xlsx")):
        try:
            app = xw.App(visible=False)
            wb = app.books.open(excel_file)
            sheet = wb.sheets[0]
            df = sheet.used_range.options(pd.DataFrame, header=1, index=False).value

            # Convertir la columna 'FechaVencimiento' a datetime, ajustar según sea necesario
            if 'FechaVencimiento' in df.columns:
                df['FechaVencimiento'] = pd.to_datetime(df['FechaVencimiento'], errors='coerce')

            wb.close()
            app.quit()

            base = os.path.basename(excel_file)
            csv_file = os.path.join(output_folder, base.replace('.xlsx', '.csv'))
            df.to_csv(csv_file, index=False, encoding='utf-8-sig', sep=';')
            print(f"Convertido {excel_file} a {csv_file}")
        except Exception as e:
            print(f"Error al convertir {excel_file} a CSV: {e}")

# Función para obtener el nombre del cliente a partir del nombre del archivo
def obtener_nombre_cliente(filename):
    base = os.path.basename(filename)
    nombre_cliente = base.split('-')[1].strip()
    return nombre_cliente

def normalizar_cuit(cuit):
    """
    Normaliza el CUIT para asegurarse de que se trate como una cadena de números enteros sin guiones ni puntos.
    """
    return str(int(cuit)).zfill(11)  # Convertir a entero para quitar decimales y asegurar 11 dígitos

datos_clientes = []

# Iterar sobre cada cliente
indice = 0

# El contador se usa para el primer cliente, ya que en este no hay que hacer click en el menú pero si en "cuenta corriente"
contador = 1
for cuit_ingresar, cuit_representado, password, posterior, anterior, cliente in zip(cuit_login_list, cuit_represent_list, password_list, posterior_list, anterior_list, clientes_list):
    # Normalizar los CUITs
    cuit_ingresar_normalizado = normalizar_cuit(cuit_ingresar)
    cuit_representado_normalizado = normalizar_cuit(cuit_representado)

    if anterior == 0:
        extraer_datos_nuevo(cuit_ingresar_normalizado, cuit_representado_normalizado, password, posterior, cliente, indice, contador)
    else:
        extraer_datos(cuit_representado_normalizado, posterior, cliente, contador)
    
    indice += 1
    contador += 1

# Convertir la lista de datos en un DataFrame de pandas
df = pd.DataFrame(datos_clientes)

# Guardar el DataFrame en un archivo Excel
archivo_salida = os.path.join(base_dir, "Data", "deudas_clientes.xlsx")
df.to_excel(archivo_salida, index=False)
